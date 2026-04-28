"""
Excel (.xlsx) report exporter with professional formatting.
"""

from __future__ import annotations

import io
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from core.models import AttendanceStatus, StudentRecord

# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_BG = "1F3864"      # Deep navy
HEADER_FG = "FFFFFF"      # White text
OK_BG = "D9EAD3"          # Soft green
SHORTAGE_BG = "FCE4D6"    # Soft red-orange
ALT_ROW_BG = "F2F7FF"     # Light blue tint
WHITE = "FFFFFF"
BORDER_COLOR = "B8C9E1"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def _apply_header_style(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=HEADER_FG, name="Arial", size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _apply_data_style(cell, row_idx: int, is_shortage: bool, center: bool = True) -> None:
    if is_shortage:
        bg = SHORTAGE_BG
    elif row_idx % 2 == 0:
        bg = ALT_ROW_BG
    else:
        bg = WHITE

    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = THIN_BORDER


def export_excel(records: List[StudentRecord], threshold: float = 75.0) -> bytes:
    """
    Generate a professional Excel report from student records.

    Args:
        records: List of StudentRecord objects.
        threshold: Attendance threshold used in this report.

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    _build_summary_sheet(wb, records, threshold)

    # ── Full report sheet ──────────────────────────────────────────────────────
    _build_report_sheet(wb, records)

    # ── Flagged students sheet ─────────────────────────────────────────────────
    flagged = [r for r in records if r.status == AttendanceStatus.SHORTAGE]
    if flagged:
        _build_flagged_sheet(wb, flagged)

    # Remove default empty sheet if others exist
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb: Workbook, records: List[StudentRecord], threshold: float) -> None:
    ws = wb.create_sheet("📊 Summary", 0)
    ws.sheet_view.showGridLines = False

    ok_count = sum(1 for r in records if r.status == AttendanceStatus.OK)
    shortage_count = len(records) - ok_count
    avg_pct = sum(r.attendance_percentage for r in records) / len(records) if records else 0

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📋 Attendance Report — Summary"
    title_cell.font = Font(bold=True, size=16, color=HEADER_BG, name="Arial")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"Threshold: {threshold}%   |   Total Students: {len(records)}"
    sub.font = Font(size=11, color="555555", name="Arial")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # KPI cards (rows 4–8)
    kpis = [
        ("Total Students", len(records), "1F3864", "FFFFFF"),
        ("✅ Meeting Threshold", ok_count, "274E13", "FFFFFF"),
        ("⚠️ Below Threshold", shortage_count, "7F0000", "FFFFFF"),
        ("Average Attendance", f"{avg_pct:.1f}%", "0B3D91", "FFFFFF"),
    ]

    headers = ["Metric", "Value"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        _apply_header_style(cell, h)

    for row_i, (label, value, bg, fg) in enumerate(kpis, start=5):
        lc = ws.cell(row=row_i, column=1, value=label)
        lc.font = Font(name="Arial", size=11, bold=True)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.font = Font(name="Arial", size=11, bold=True, color=fg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = THIN_BORDER

        vc = ws.cell(row=row_i, column=2, value=value)
        vc.font = Font(name="Arial", size=12, bold=True, color=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = THIN_BORDER

        ws.row_dimensions[row_i].height = 22

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 5


def _build_report_sheet(wb: Workbook, records: List[StudentRecord]) -> None:
    ws = wb.create_sheet("📋 Full Report")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Student Name", 26, False),
        ("Roll Number", 14, True),
        ("Total Classes", 14, True),
        ("Classes Attended", 16, True),
        ("Attendance %", 14, True),
        ("Deficit Classes", 15, True),
        ("Status", 14, True),
    ]

    # Header row
    for col_idx, (col_name, width, _) in enumerate(columns, start=1):
        _apply_header_style(ws.cell(row=1, column=col_idx), col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_i, record in enumerate(records, start=2):
        is_shortage = record.status == AttendanceStatus.SHORTAGE
        values = [
            record.name,
            record.roll_number,
            record.total_classes,
            record.classes_attended,
            f"{record.attendance_percentage:.1f}%",
            record.deficit_classes,
            record.status.value,
        ]
        for col_idx, (val, (_, _, center)) in enumerate(zip(values, columns), start=1):
            cell = ws.cell(row=row_i, column=col_idx, value=val)
            _apply_data_style(cell, row_i, is_shortage, center=center)
        ws.row_dimensions[row_i].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _build_flagged_sheet(wb: Workbook, flagged: List[StudentRecord]) -> None:
    ws = wb.create_sheet("⚠️ Flagged Students")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Student Name", 26),
        ("Roll Number", 14),
        ("Attendance %", 14),
        ("Deficit Classes", 16),
        ("AI Warning Message", 60),
    ]

    for col_idx, (col_name, width) in enumerate(columns, start=1):
        _apply_header_style(ws.cell(row=1, column=col_idx), col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    for row_i, record in enumerate(flagged, start=2):
        values = [
            record.name,
            record.roll_number,
            f"{record.attendance_percentage:.1f}%",
            record.deficit_classes,
            record.ai_warning or "—",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_idx, value=val)
            center = col_idx != 1 and col_idx != 5
            _apply_data_style(cell, row_i, is_shortage=True, center=center)
        ws.row_dimensions[row_i].height = 50

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
